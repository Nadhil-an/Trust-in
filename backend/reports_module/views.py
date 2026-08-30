"""
Reports Module — PDF and Excel generation for all modules.
Uses ReportLab for PDF and openpyxl for Excel.
"""
import io
from datetime import datetime
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from core.permissions import IsAnyStaff


def get_common_filters(request):
    return {
        'from_date': request.query_params.get('from_date'),
        'to_date': request.query_params.get('to_date'),
        'status': request.query_params.get('status'),
        'category': request.query_params.get('category'),
    }


def build_pdf_response(title: str, headers: list, rows: list, filename: str) -> HttpResponse:
    """Build a styled PDF using ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Header
    header_style = ParagraphStyle('header', parent=styles['Heading1'],
                                   fontSize=16, textColor=colors.HexColor('#0A2463'),
                                   alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph("Sree Lakshmi Charitable Trust", header_style))
    elements.append(Paragraph(title, ParagraphStyle('title', parent=styles['Heading2'],
                                                     fontSize=13, alignment=TA_CENTER,
                                                     textColor=colors.HexColor('#1E4DB7'), spaceAfter=2)))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}", sub_style))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1E4DB7')))
    elements.append(Spacer(1, 0.3*cm))

    # Table
    data = [headers] + rows
    col_count = len(headers)
    col_width = (26*cm) / col_count
    table = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E4DB7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#DBEAFE')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#94a3b8')),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_excel_response(title: str, headers: list, rows: list, filename: str) -> HttpResponse:
    """Build an Excel workbook using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    blue = '1E4DB7'
    light_blue = 'DBEAFE'
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color=blue, end_color=blue, fill_type='solid')
    row_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(left=Side(style='thin', color='CBD5E1'),
                  right=Side(style='thin', color='CBD5E1'),
                  top=Side(style='thin', color='CBD5E1'),
                  bottom=Side(style='thin', color='CBD5E1'))

    # Title row
    ws.merge_cells(f'A1:{chr(64+len(headers))}1')
    ws['A1'] = f"Sree Lakshmi Charitable Trust — {title}"
    ws['A1'].font = Font(bold=True, size=13, color=blue)
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{chr(64+len(headers))}2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].alignment = center

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    # Data
    for r_idx, row in enumerate(rows, 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx % 2 == 0:
                cell.fill = row_fill
            cell.border = thin
            cell.alignment = Alignment(vertical='center')

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[4].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Report Views ──────────────────────────────────────────────────

class AssessmentRequestReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        f = get_common_filters(request)
        from manager_module.models import AssessmentRequest
        qs = AssessmentRequest.objects.select_related('requested_by', 'reviewed_by').all()
        if f['from_date']: qs = qs.filter(created_at__date__gte=f['from_date'])
        if f['to_date']: qs = qs.filter(created_at__date__lte=f['to_date'])
        if f['status']: qs = qs.filter(status=f['status'])
        if f['category']: qs = qs.filter(category=f['category'])

        headers = ['Request No', 'Date', 'Purpose', 'Category', 'Amount Requested',
                   'Amount Approved', 'Status', 'Requested By', 'Approved By']
        rows = [[r.request_number, str(r.created_at.date()), r.purpose, r.get_category_display(),
                 f'₹{r.amount_requested:,.2f}',
                 f'₹{r.amount_approved:,.2f}' if r.amount_approved else '-',
                 r.get_status_display(),
                 r.requested_by.full_name if r.requested_by else '',
                 r.reviewed_by.full_name if r.reviewed_by else ''] for r in qs]

        if fmt == 'pdf':
            return build_pdf_response('Assessment Requests Report', headers, rows,
                                       f'requests_report_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Requests Report', headers, rows,
                                         f'requests_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows})


class CashBookReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        from accounts_module.models import CashTransaction
        qs = CashTransaction.objects.select_related('cash_account').all()
        if from_date: qs = qs.filter(date__gte=from_date)
        if to_date: qs = qs.filter(date__lte=to_date)

        headers = ['Date', 'Reference', 'Description', 'Type', 'Receipt (₹)', 'Payment (₹)', 'Balance (₹)']
        rows = [[str(t.date), t.reference_id, t.description, t.get_transaction_type_display(),
                 f'{t.amount:,.2f}' if t.transaction_type in ['RECEIPT', 'TRANSFER_IN', 'OPENING'] else '',
                 f'{t.amount:,.2f}' if t.transaction_type not in ['RECEIPT', 'TRANSFER_IN', 'OPENING'] else '',
                 f'{t.balance_after:,.2f}'] for t in qs]

        if fmt == 'pdf':
            return build_pdf_response('Cash Book', headers, rows,
                                       f'cashbook_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Cash Book', headers, rows,
                                         f'cashbook_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows})


class IncomeReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        f = get_common_filters(request)
        from accounts_module.models import Income
        qs = Income.objects.all()
        if f['from_date']: qs = qs.filter(date__gte=f['from_date'])
        if f['to_date']: qs = qs.filter(date__lte=f['to_date'])
        if f['category']: qs = qs.filter(source=f['category'])

        headers = ['Receipt No', 'Date', 'Donor', 'Source', 'Amount (₹)', 'Payment Method', 'Account']
        rows = [[i.receipt_number, str(i.date), i.donor_name, i.get_source_display(),
                 f'{i.amount:,.2f}', i.get_payment_method_display(), i.account_type] for i in qs]

        if fmt == 'pdf':
            return build_pdf_response('Income Report', headers, rows,
                                       f'income_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Income Report', headers, rows,
                                         f'income_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows,
                         'total': float(sum(r.amount for r in qs))})


class ExpenseReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        f = get_common_filters(request)
        from accounts_module.models import Expense
        qs = Expense.objects.all()
        if f['from_date']: qs = qs.filter(date__gte=f['from_date'])
        if f['to_date']: qs = qs.filter(date__lte=f['to_date'])
        if f['category']: qs = qs.filter(category=f['category'])
        if f['status']: qs = qs.filter(status=f['status'])

        headers = ['Expense ID', 'Date', 'Payee', 'Category', 'Amount (₹)', 'Payment Method', 'Status']
        rows = [[e.expense_id, str(e.date), e.payee, e.get_category_display(),
                 f'{e.amount:,.2f}', e.get_payment_method_display(), e.status] for e in qs]

        if fmt == 'pdf':
            return build_pdf_response('Expense Report', headers, rows,
                                       f'expense_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Expense Report', headers, rows,
                                         f'expense_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows,
                         'total': float(sum(r.amount for r in qs))})


class MemberReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        f = get_common_filters(request)
        from hr_module.models import Member
        qs = Member.objects.all()
        if f['status']: qs = qs.filter(status=f['status'])

        headers = ['Member ID', 'Name', 'Phone', 'Email', 'Type', 'Joining Date', 'Status']
        rows = [[m.member_id, m.full_name, m.phone, m.email,
                 m.get_membership_type_display(), str(m.joining_date),
                 m.get_status_display()] for m in qs]

        if fmt == 'pdf':
            return build_pdf_response('Member Report', headers, rows,
                                       f'members_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Member Report', headers, rows,
                                         f'members_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows})


class PayrollReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        from hr_module.models import MonthlyPayroll
        qs = MonthlyPayroll.objects.select_related('employee').all()
        if month: qs = qs.filter(month=month)
        if year: qs = qs.filter(year=year)

        headers = ['Payroll ID', 'Employee', 'Month/Year', 'Basic', 'Gross', 'Deductions', 'Net', 'Status']
        rows = [[p.payroll_id, p.employee.full_name, f'{p.month}/{p.year}',
                 f'₹{p.basic_salary:,.2f}', f'₹{p.gross_salary:,.2f}',
                 f'₹{p.pf_deduction + p.other_deductions:,.2f}',
                 f'₹{p.net_salary:,.2f}', p.get_status_display()] for p in qs]

        if fmt == 'pdf':
            return build_pdf_response('Payroll Report', headers, rows,
                                       f'payroll_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Payroll Report', headers, rows,
                                         f'payroll_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows})


class TransactionReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        fmt = request.query_params.get('format', 'json')
        f = get_common_filters(request)
        from accounts_module.models import Transaction
        qs = Transaction.objects.all()
        if f['from_date']: qs = qs.filter(date__gte=f['from_date'])
        if f['to_date']: qs = qs.filter(date__lte=f['to_date'])
        if f['category']: qs = qs.filter(transaction_type=f['category'])

        headers = ['Txn ID', 'Date', 'Type', 'Description', 'Debit (₹)', 'Credit (₹)', 'Account', 'Ref']
        rows = [[t.transaction_id, str(t.date), t.get_transaction_type_display(), t.description,
                 f'{t.debit:,.2f}' if t.debit else '',
                 f'{t.credit:,.2f}' if t.credit else '',
                 t.account_type, t.reference_id] for t in qs]

        if fmt == 'pdf':
            return build_pdf_response('Transaction Report', headers, rows,
                                       f'transactions_{datetime.now().strftime("%Y%m%d")}.pdf')
        if fmt == 'excel':
            return build_excel_response('Transaction Report', headers, rows,
                                         f'transactions_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return Response({'count': qs.count(), 'headers': headers, 'rows': rows})


class StaffPerformanceReportView(APIView):
    permission_classes = [IsAnyStaff]

    def get(self, request):
        from datetime import datetime, timedelta
        from django.db.models import Sum, Q
        from core.models import User, Role
        from hr_module.models import Member
        from accounts_module.models import Income
        from manager_module.models import AssessmentRequest

        period = request.query_params.get('period', 'weekly').lower()
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')

        # Parse custom dates if provided, otherwise use default period duration
        today = timezone.now().date()
        start_date = None
        end_date = None

        if from_date_str:
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    start_date = datetime.strptime(from_date_str, fmt).date()
                    break
                except ValueError:
                    pass

        if to_date_str:
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    end_date = datetime.strptime(to_date_str, fmt).date()
                    break
                except ValueError:
                    pass

        if not end_date:
            end_date = today
        if not start_date:
            days = 7 if period == 'weekly' else 30
            start_date = end_date - timedelta(days=days)

        duration_days = max(1, (end_date - start_date).days)
        prev_end_date = start_date - timedelta(days=1)
        prev_start_date = prev_end_date - timedelta(days=duration_days)

        # Get staff members (Users ONLY with role = STAFF or mobile field staff roles)
        staff_users = User.objects.filter(
            Q(role=Role.STAFF) |
            Q(role=Role.FIELD_ASSESSMENT_OFFICER) |
            Q(role=Role.ASSESSMENT_CALCULATION_OFFICER) |
            Q(role=Role.GENERAL_ENQUIRY_OFFICER),
            is_active=True
        ).order_by('full_name')

        performance_list = []
        total_donations_all = 0.0
        total_leads_all = 0
        total_enquiries_all = 0

        for user in staff_users:
            # 1. Total Donation created by user
            curr_donations = Income.objects.filter(
                created_by=user,
                source='DONATION',
                date__gte=start_date,
                date__lte=end_date
            ).aggregate(total=Sum('amount'))['total'] or 0.0

            prev_donations = Income.objects.filter(
                created_by=user,
                source='DONATION',
                date__gte=prev_start_date,
                date__lte=prev_end_date
            ).aggregate(total=Sum('amount'))['total'] or 0.0

            curr_donations = float(curr_donations)
            prev_donations = float(prev_donations)

            if prev_donations > 0:
                growth_pct = round(((curr_donations - prev_donations) / prev_donations) * 100, 1)
                growth_str = f"+{growth_pct}%" if growth_pct >= 0 else f"{growth_pct}%"
            else:
                growth_str = "+10%" if curr_donations > 0 else "0%"

            # 2. Total Leads (Assessment requests created by user + Members created by user)
            leads_requests = AssessmentRequest.objects.filter(
                requested_by=user,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).count()

            leads_members = Member.objects.filter(
                created_by=user,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).count()

            total_leads = leads_requests + leads_members

            # 3. Total Enquiries (Assigned to user or requested/reviewed by user)
            enquiries_count = AssessmentRequest.objects.filter(
                Q(assigned_fao=user) | Q(assigned_aco=user) | Q(assigned_geo=user) | Q(reviewed_by=user) | Q(requested_by=user),
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).distinct().count()

            engagement_rate = f"{round((total_leads / (enquiries_count or 1)) * 100, 1)}%" if enquiries_count > 0 else "0%"
            month_ratio = total_leads if total_leads > 0 else enquiries_count

            total_donations_all += curr_donations
            total_leads_all += total_leads
            total_enquiries_all += enquiries_count

            display_name = user.full_name or user.username

            performance_list.append({
                'id': str(user.id),
                'name': display_name,
                'role': user.get_role_display(),
                'donation': curr_donations,
                'growth': growth_str,
                'leads': total_leads,
                'enquiries': enquiries_count,
                'engagement_rate': engagement_rate,
                'month_ratio': month_ratio,
            })

        performance_list.sort(key=lambda x: (x['donation'], x['leads']), reverse=True)

        return Response({
            'period': period,
            'from_date': start_date.strftime('%d %b %Y').upper(),
            'to_date': end_date.strftime('%d %b %Y').upper(),
            'prepared_by': 'HR And Accounts',
            'summary': {
                'total_donation': total_donations_all,
                'total_leads': total_leads_all,
                'total_enquiries': total_enquiries_all,
            },
            'staff_performance': performance_list,
        })

